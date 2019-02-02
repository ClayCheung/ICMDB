from openpyxl import load_workbook
import re


def getAllDeviceList(tableFilePath):
    """
    获取所有设备的设备名列表
    :param tableFilePath:
    :return:
    """
    wb = load_workbook(tableFilePath)

    # 打开第一个sheet
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])

    DeviceSet = set([])

    for cell in sheet['B']:
        if cell.value != None and cell.value.startswith("ZJ"):
            DeviceSet.add(cell.value)

    for cell in sheet['E']:
        if cell.value != None and cell.value.startswith("ZJ"):
            DeviceSet.add(cell.value)

    # print(DeviceSet)

    return DeviceSet


def getNetDeviceList(tableFilePath):
    """
    获取网络设备的设备名列表，并且排序
    :param tableFilePath:
    :return:
    """
    DeviceSet = getAllDeviceList(tableFilePath)

    # 检测是否为网络设备,把所有网络设备传入一个新set

    netDeviceSet = set([])

    for device in DeviceSet:
        if device.find('-SRV') == -1 and device.find('-DA') == -1 \
                and device.find('-DAC') == -1 and device.find('-SDS') == -1:
            netDeviceSet.add(device)


    netDeviceSet = sorted(netDeviceSet,
                           key=lambda netDevList: (netDevList.split('-')[5], netDevList.split('-')[6]))
    # print(netDeviceSet)
    return netDeviceSet


def getServerDeviceList(tableFilePath):
    """
    获取网络设备的设备名列表，并且排序
    :param tableFilePath:
    :return:
    """

    DeviceSet = getAllDeviceList(tableFilePath)

    # 检测是否为服务器,把所有服务器传入一个新set

    serverDeviceSet = set([])

    for device in DeviceSet:
        if device.find('-SRV') != -1 :
            serverDeviceSet.add(device)
        elif device.find('-SDS') != -1 :
            serverDeviceSet.add(device)

    serverDeviceSet = sorted(serverDeviceSet,
                             key=lambda netDevList: (netDevList.split('-')[5], netDevList.split('-')[6]))

    return serverDeviceSet


def getStorDeviceList(tableFilePath):
    """
    获取存储设备的设备名列表，并且排序
    :param tableFilePath:
    :return:
    """

    DeviceSet = getAllDeviceList(tableFilePath)

    # 检测是否为存储设备,把所有服务器传入一个新set

    storDeviceSet = set([])

    for device in DeviceSet:
        if device.find('-DA') != -1 or device.find('-DAC') != -1:
            storDeviceSet.add(device)

    storDeviceSet = sorted(storDeviceSet,
                           key=lambda netDevList: (netDevList.split('-')[5], netDevList.split('-')[6]))

    return storDeviceSet

def mapSubTypeId(e_type_str):
    subTypeId = {
        '交换机': 0,
        '防火墙': 1,
        '负载均衡器': 2,
        'SAN交换机': 3,
        '路由器': 4
    }
    return subTypeId[e_type_str]

def mapSubType(e_type_id):
    subType = {
        0: '交换机',
        1: '防火墙',
        2: '负载均衡器',
        3: 'SAN交换机',
        4: '路由器'
    }
    return subType[int(e_type_id)]

def mapVender(e_type_id, e_vender_id):
    vender = {
        0: {# 交换机
            0: 'H3C',
            1: '华为',
            2: '锐捷',
            3: '思科',
        },
        1: {# 防火墙
            0: 'H3C',
            1: '华为',
        },
        2: {  # 负载均衡器
            0: '迪普',
            1: 'Array',
            2: 'F5',
        },
        3: {  # SAN交换机
            0: '博科',
            1: '思科',
        },
        4: {  # 路由器
            0: 'H3C',
            1: '华为',
            2: '思科',
        },
    }

    return vender[int(e_type_id)][int(e_vender_id)]


def mapModel(e_type_id, e_vender_id, e_model_id):

    model = {
        0: {# 交换机
            0: {# H3C
                0: 'S6900-4F',
                1: 'S6800',
                2: 'S12500',
                3: 'S7600',
                4: 'S6800',
            },
            1: {# 华为
                0: 'CE12800',
                1: 'CE6851',
                2: 'S9300',
            },
            2: {  # 锐捷
                0: 'RJ5500',
            },
            3: {  # 思科
                0: 'N7710',
                1: 'N5K',
            },
        },
        1: {# 防火墙
            0: {  # H3C
                0: 'M9000',
            },
            1: {  # 华为
                0: 'E8000E-X8',
            },
        },
        2: {  # 负载均衡器
            0: {  # 迪普
                0: 'DPtech ADX3000',
            },
            1: {  # Array
                0: 'APV 9650',
            },
            2: {  # F5
                0: 'F5负载均衡',
            },
        },
        3: {  # SAN交换机
            0: {  # 博科
                0: 'Brocade 6510',
            },
            1: {  # 思科
                0: '思科SAN交换机',
            },
        },
        4: {  # 路由器
            0: {  # H3C
                0: 'H3C路由器',
            },
            1: {  # 华为
                0: '华为路由器',
            },
            2: {  # 思科
                0: '思科路由器',
            },
        },
    }

    return model[int(e_type_id)][int(e_vender_id)][int(e_model_id)]


def get_cabinetNum(devName, cableTableFile_path):
    """
    输入资产名称和布线表，输出资产的机柜号
    :param devName:
    :param cableTableFile_path:
    :return:
    """

    wb = load_workbook(cableTableFile_path)

    # 打开第一个sheet
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])

    cabinetNum = ''
    for cell in sheet['B']:
        if cell.value != None and cell.value.strip() == devName:
            cabinetNum = sheet['C'][cell.row-1].value
            break
    for cell in sheet['E']:
        if cell.value != None and cell.value.strip() == devName:
            cabinetNum = sheet['F'][cell.row-1].value
            break

    return cabinetNum

def parse_cabinetNum(cabinetNum):
    """
    例子：输入 'F24' ,输出 ('F','24')
    :param cabinetNum:
    :return:
    """
    cabinetNum = cabinetNum.strip()

    if len(cabinetNum)!=3:
        return -1
    else:
        return (cabinetNum[0],cabinetNum[1:3])


def parse_serverVender(devName):
    """
    从服务器的设备名解析出服务器的厂商
    :param devName:
    :return:
    vender: 厂商名
    -1: 服务器命名不合规
    """
    fragment = devName.strip().split('-')
    if len(fragment) >= 6:
        model = fragment[5]
        if model.find('2288')!=-1 or model.upper().find('HW')!=-1:# 型号中带2288或HW
            vender = '华为'
        elif model.find('380')!=-1:# 型号中带380
            vender= '新华三'
        elif model.upper().find('DELL')!=-1 or model.find('730')!=-1:
            vender= '戴尔'
        else:
            vender= '未知厂商'
    else:# 服务器命名不合规
        return -1
    return vender


def parse_serverModel(devName):
    """
    从服务器的设备名解析出服务器的型号
    :param devName:
    :return:
    model: 设备型号
    -1: 服务器命名不合规
    """
    fragment = devName.strip().split('-')
    if len(fragment) >= 6:
        model = fragment[5]
        if model.find('2288')!=-1 :# 型号中带2288或HW
            model_s = 'RH2288H'
        elif model.find('380')!=-1:# 型号中带380
            model_s= 'DL380'
        elif  model.find('730')!=-1:
            model_s= 'R730'
        else:
            model_s= '未知厂商'
    else:# 服务器命名不合规
        return -1
    return model_s


def parse_storVender(devName):
    """
    从存储设备的设备名解析出存储的厂商
    :param devName:
    :return:
    vender: 厂商名
    -1: 存储设备命名不合规
    """
    fragment = devName.strip().split('-')
    if len(fragment) >= 6:
        model = fragment[5]
        if model.find('6800')!=-1 or model.upper().find('HW')!=-1:# 型号中带6800或HW
            vender = '华为'
        else:
            vender= '未知厂商'
    else:# 服务器命名不合规
        return -1
    return vender


def parse_storModel(devName):
    """
    从存储设备的设备名解析出存储的型号
    :param devName:
    :return:
    model: 设备型号
    -1: 存储设备命名不合规
    """
    fragment = devName.strip().split('-')
    if len(fragment) >= 6:
        model = fragment[5]
        if model.find('6800')!=-1 :# 型号中带6800
            model_s = 'OceanStor 6800 V3'
        elif model.find('5600')!=-1:# 型号中带5600
            model_s= 'OceanStor 5600'
        else:
            model_s= '未知厂商'
    else:# 服务器命名不合规
        return -1
    return model_s


def parse_machine_room(devName):
    """
    从资产名称中 解析出 机房地区、机房名
    :param devName:
    :return:
    (机房地区, 机房名)
    """
    fragment = devName.strip().split('-')
    if len(fragment) >= 6:
        region = fragment[2]
        machine_room = fragment[3]
    return (region, machine_room)


def check_cabinetNum(tableFilePath):
    """
    检测机柜号是否和资产名中解析出的机柜号匹配
    :param tableFilePath:
    :return:
    不匹配的错误日志
    """
    wb = load_workbook(tableFilePath)

    # 打开第一个sheet
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])


    err_cabinetNum_log = []
    for cell in sheet['B']:
        if cell.value != None and cell.value.startswith("ZJ"):
            # print('行号！！！！！！！！')
            # print(cell.row)
            # print(sheet['C'][cell.row].value)
            cabinetNum_parse = cell.value.strip().split('-')[4]
            cabinetNum = sheet['C'][cell.row-1].value.strip()
            if cabinetNum != cabinetNum_parse:
                err_row = []
                for c in sheet[cell.row]:
                    err_row.append(c.value)
                log = "第{0}行 资产名称与机柜号不匹配 ： {1}".format(str(cell.row), err_row)
                err_cabinetNum_log.append(log)

    for cell in sheet['E']:
        if cell.value != None and cell.value.startswith("ZJ"):
            cabinetNum_parse = cell.value.strip().split('-')[4]
            cabinetNum = sheet['F'][cell.row-1].value.strip()
            if cabinetNum != cabinetNum_parse:
                err_row = []
                for c in sheet[cell.row]:
                    err_row.append(c.value)
                log = "第{0}行 资产名称与机柜号不匹配 ： {1}".format(str(cell.row), err_row)
                err_cabinetNum_log.append(log)

    if len(err_cabinetNum_log)==0 :
        err_cabinetNum_log.append('机柜号无错误！')

    return err_cabinetNum_log


def parse_port_H3CS6900(port_str):
    """
    解析S6900类型的接口，即 speed chassis/slot/portNum
    :param port_str:
    :return:
    [int:speed, int:chassis, int:slot, str:portNum]
    """
    speed_map = {
        '': 1,
        '1': 1,
        'X': 10,
        '10': 10,
        '40': 40,
        '100': 100,
    }
    ret = re.match('(.*)G[E]?\s*([1-2])/([1-4])/(\d{1,2})', port_str)
    if ret == None:
        if port_str.upper()=='MGMT':
            return [None, None, None, 'mgmt']
        else:
            return None
    else:
        if int(ret.group(4))>=1 and int(ret.group(4))<=26:
            return [speed_map[ret.group(1)], int(ret.group(2)), int(ret.group(3)), ret.group(4)]
        else:
            return None

def parse_port_H3CS6800(port_str):
    """
    解析S6800类型的接口，即 speed chassis/slot/portNum
    :param port_str:
    :return:
    [int:speed, int:chassis, int:slot, str:portNum]
    """
    speed_map = {
        '': 1,
        '1': 1,
        'X': 10,
        '10': 10,
        '40': 40,
        '100': 100,
    }
    ret = re.match('(.*)G[E]?\s*([1-2])/([1-4])/(\d{1,2})', port_str)
    if ret == None:
        if port_str.upper()=='MGMT':
            return [None, None, None, 'mgmt']
        else:
            return None
    else:
        if int(ret.group(4))>=1 and int(ret.group(4))<=54:
            return [speed_map[ret.group(1)], int(ret.group(2)), int(ret.group(3)), ret.group(4)]
        else:
            return None


def parse_port_H3CS12500(port_str):
    """
    解析S12500类型的接口(12500 7600系列)，即 speed chassis/slot/subSlot/portNum
    :param port_str:
    :return:
    [int:speed, int:chassis, int:slot, int:subSlot, str:portNum]
    """
    speed_map = {
        '': 1,
        '1': 1,
        'X': 10,
        '10': 10,
        '40': 40,
        '100': 100,
    }
    ret = re.match('(.*)G[E]?\s*([1-4])/(\d{1,2})/([0-3])/(\d{1,2})', port_str)
    if ret == None:# 没有匹配上
        if port_str.upper()=='MGMT':
            return [None, None, None, None, 'mgmt']
        else:
            return None
    else:
        if int(ret.group(5))>=1 and int(ret.group(5))<=48 and int(ret.group(3))>=0 and int(ret.group(3))<=16:
            return  [speed_map[ret.group(1)], int(ret.group(2)), int(ret.group(3)), int(ret.group(4)), ret.group(5)]
        else:
            return None


def parse_port_HWS9300(port_str):
    """
    解析huawei S9300类型的接口，即 speed chassis/slot/subSlot/portNum
    :param port_str:
    :return:
    [int:speed, int:chassis, int:slot, int:subSlot, str:portNum]
    """
    speed_map = {
        '': 1,
        '1': 1,
        'X': 10,
        '10': 10,
        '40': 40,
        '100': 100,
    }
    ret = re.match('(.*)G[E]?\s*([1-4])/(\d{1,2})/([0-3])/(\d{1,2})', port_str)
    if ret == None:  # 没有匹配上
        if port_str.upper()=='MGMT':
            return [None, None, None, None, 'mgmt']
        else:
            return None
    else:
        if int(ret.group(5))>=0 and int(ret.group(5))<=47 and int(ret.group(3))>=0 and int(ret.group(3))<=16:
            return  [speed_map[ret.group(1)], int(ret.group(2)), int(ret.group(3)), int(ret.group(4)), ret.group(5)]
        else:
            return None



def parse_port_Brocade6510(port_str):
    """
    解析Brocade 6510类型的接口，即 port slot/portNum
    :param port_str:
    :return:
    [int:slot, str:portNum]
    """

    ret = re.match('PORT\s*([1-4])/(\d{1,2})', port_str.upper())
    if ret == None:  # 没有匹配上
        if port_str.upper()=='MGMT':
            return [None, 'mgmt']
        else:
            return None
    else:
        return [int(ret.group(1)), ret.group(2)]



def parse_port_E8000E(port_str):
    """
    解析huawei FW E8000E类型的接口，即 speed slot/subSlot/portNum
    :param port_str:
    :return:
    [int:speed, int:slot, int:subSlot, str:portNum]
    """
    speed_map = {
        '': 1,
        '1': 1,
        'X': 10,
        '10': 10,
        '40': 40,
        '100': 100,
    }
    ret = re.match('(.*)G[E]?\s*([1-8])/([0-3])/(\d{1,2})', port_str)
    if ret == None:  # 没有匹配上
        if port_str.upper() == 'MGMT':
            return [None, None, None, 'mgmt']
        else:
            return None
    else:
        return [speed_map[ret.group(1)], int(ret.group(2)), int(ret.group(3)), ret.group(4)]