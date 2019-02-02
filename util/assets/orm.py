from assets.models import Asset, Locate , Server, StorDevice ,Port
from util.cableTable.tools import *

from openpyxl import load_workbook




def update_or_create_cabinetNum(devName, cabinetNum):
    """
    输入资产名称和机柜号，检测是否存在是这个资产名的资产，如果有则更新机柜号
    如果没有，则创建该资产并更新设备名、机柜号
    :param devName:
    :param cabinetNum:
    :return:
    延用update_or_create()的 return
    返回一个元祖object是创建或更新的对象，created是布尔值 (object, created)
    """


    locateRet = Locate.objects.update_or_create(region=parse_machine_room(devName)[0],
                                    machine_room=parse_machine_room(devName)[1],
                                    cabinet=parse_cabinetNum(cabinetNum)[0],
                                    cabinet_num=parse_cabinetNum(cabinetNum)[1],)
    locateObj = locateRet[0]

    ret = Asset.objects.update_or_create(name=devName,
                                   defaults={
                                       'locate': locateObj,

                                   }
    )

    return ret


def update_or_create_vender(devName, vender):
    """
     输入资产名称和设备厂商名，检测是否存在是这个资产名的资产，如果有则更新设备厂商
    如果没有，则创建该资产并更新设备名、设备厂商
    :param devName:
    :param vender:
    :return:
    延用update_or_create()的 return
    返回一个元祖object是创建或更新的对象，created是布尔值 (object, created)
    """
    ret = Asset.objects.update_or_create(name=devName,
                                   defaults={
                                       'vender': vender,
                                   }
    )
    return ret


def update_or_create_model(devName, model):
    """
     输入资产名称和设备型号，检测是否存在是这个资产名的资产，如果有则更新设备型号
    如果没有，则创建该资产并更新设备名、设备型号
    :param devName:
    :param model:
    :return:
    延用update_or_create()的 return
    返回一个元祖object是创建或更新的对象，created是布尔值 (object, created)
    """
    ret = Asset.objects.update_or_create(name=devName,
                                   defaults={
                                       'model': model,
                                   }
    )
    return ret


def update_or_create_assetType(devName, assetType):
    """
    输入资产名称和资产类型，检测是否存在是这个资产名的资产，如果有则更新资产类型
    如果没有，则创建该资产并更新设备名、资产类型
    :param devName:
    :param assetType:
    :return:
    延用update_or_create()的 return
    返回一个元祖object是创建或更新的对象，created是布尔值 (object, created)
    """
    ret = Asset.objects.update_or_create(name=devName,
                                   defaults={
                                       'asset_type': assetType
                                   }
    )
    return ret


def map_ServerToAsset(devName):
    """
    创建一个Server对象,映射OneToOne到 名为devName的Asset对象
    :param devName:
    :return:
    """
    assetObj = Asset.objects.filter(name=devName)[0]
    ret = Server.objects.update_or_create(asset=assetObj)
    return ret


def map_StorDeviceToAsset(devName):
    """
    创建一个StorDevice对象,映射OneToOne到 名为devName的Asset对象
    :param devName:
    :return:
    """
    assetObj = Asset.objects.filter(name=devName)[0]
    ret = StorDevice.objects.update_or_create(asset=assetObj)
    return ret


def checkPort_and_mapToAsset(tableFilePath):
    """
    检查端口格式，
    正确：传入数据库
    不正确：返回日志
    :param tableFilePath:
    :return:
    """
    wb = load_workbook(tableFilePath)

    # 打开第一个sheet
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])

    err_portFormat_log = []

    def doThis(devName_col, port_col):
        for cell in sheet[devName_col]:
            if cell.value != None and cell.value.startswith("ZJ"):
                port = sheet[port_col][cell.row-1].value.strip()
                devName = cell.value.strip()
                asset = Asset.objects.filter(name=devName)[0]
                model = asset.model
                if model=='S6900-4F':
                    res = parse_port_H3CS6900(port)
                    if res != None:# 返回不是NONE，说明端口格式正确，则把端口信息传入数据库
                        # print(res)
                        try:
                            Port.objects.create(asset=asset,
                                                speed=res[0],
                                                chassisNum=res[1],
                                                slotNum=res[2],
                                                portNum=res[3])
                        except Exception as e:
                            print(e)
                            print("有重复的端口")

                    else:# 返回值为NONE，说明端口格式有误
                        log = "第{0}行 设备<{1}>的端口 {2} 格式错误".format(str(cell.row), devName, port)
                        err_portFormat_log.append(log)
                elif model=='S12500' or model=='S7600':
                    res = parse_port_H3CS12500(port)
                    if res != None:# 返回不是NONE，说明端口格式正确，则把端口信息传入数据库
                        try:
                            Port.objects.create(asset=asset,
                                                speed=res[0],
                                                chassisNum=res[1],
                                                slotNum=res[2],
                                                subSlotNum=res[3],
                                                portNum=res[4])
                        except Exception as e:
                            print(e)
                            print("有重复的端口")
                    else:# 返回值为NONE，说明端口格式有误
                        log = "第{0}行 设备<{1}>的端口 {2} 格式错误".format(str(cell.row), devName, port)
                        err_portFormat_log.append(log)


                elif model == 'S9300':
                    res = parse_port_HWS9300(port)
                    if res != None:  # 返回不是NONE，说明端口格式正确，则把端口信息传入数据库
                        try:
                            Port.objects.create(asset=asset,
                                                speed=res[0],
                                                chassisNum=res[1],
                                                slotNum=res[2],
                                                subSlotNum=res[3],
                                                portNum=res[4])
                        except Exception as e:
                            print(e)
                            print("有重复的端口")
                    else:  # 返回值为NONE，说明端口格式有误
                        log = "第{0}行 设备<{1}>的端口 {2} 格式错误".format(str(cell.row), devName, port)
                        err_portFormat_log.append(log)


                elif model == 'Brocade 6510':
                    res = parse_port_Brocade6510(port)
                    if res != None:  # 返回不是NONE，说明端口格式正确，则把端口信息传入数据库
                        try:
                            Port.objects.create(asset=asset,
                                                slotNum=res[0],
                                                portNum=res[1])
                        except Exception as e:
                            print(e)
                            print("有重复的端口")
                    else:  # 返回值为NONE，说明端口格式有误
                        log = "第{0}行 设备<{1}>的端口 {2} 格式错误".format(str(cell.row), devName, port)
                        err_portFormat_log.append(log)


                elif model == 'E8000E-X8':
                    res = parse_port_E8000E(port)
                    if res != None:  # 返回不是NONE，说明端口格式正确，则把端口信息传入数据库
                        try:
                            Port.objects.create(asset=asset,
                                                speed=res[0],
                                                slotNum=res[1],
                                                subSlotNum=res[2],
                                                portNum=res[3])
                        except Exception as e:
                            print(e)
                            print("有重复的端口")
                    else:  # 返回值为NONE，说明端口格式有误
                        log = "第{0}行 设备<{1}>的端口 {2} 格式错误".format(str(cell.row), devName, port)
                        err_portFormat_log.append(log)




                elif asset.asset_type=='server' or asset.asset_type=='storagedevice':# 如果是服务器或者存储设备，则不检查格式之间存入数据库
                    try:
                        Port.objects.create(asset=asset,
                                            portNum=port)
                    except Exception as e:
                        print(e)
                        print("有重复的端口")

                else: # 如果是没有定义 parse_port方法的设备，则按照服务器来存入port数据库
                    try:
                        Port.objects.create(asset=asset,
                                            portNum=port)
                    except Exception as e:
                        print(e)
                        print("有重复的端口")



    doThis('B', 'D')
    doThis('E', 'G')



    if len(err_portFormat_log)==0:
        err_portFormat_log.append('网络设备端口格式无错误！')

    return err_portFormat_log


def get_portObj(sheet, devName_col, port_col, row):
    """
    输入 sheet 设备名所在列 端口所在列 所在行，
    定位到一个已经录入数据库的 Port Obj
    :param sheet: pyopenxl sheet
    :param devName_col: str
    :param port_col: str
    :param row: int
    :return:
    返回 Port Obj 或者
    没找到 返回None
    """
    port = sheet[port_col][row - 1].value.strip()
    devName = sheet[devName_col][row - 1].value.strip()
    asset = Asset.objects.filter(name=devName)[0]
    model = asset.model

    if model == 'S6900-4F':
        res = parse_port_H3CS6900(port)
        if res != None:  # 返回不是NONE，说明端口格式正确, 则 定位到该 Port Obj

            portObj = Port.objects.filter(asset=asset,
                                          speed=res[0],
                                          chassisNum=res[1],
                                          slotNum=res[2],
                                          portNum=res[3])[0]
        else:  # 返回值为NONE，说明端口格式有误
            return None

    elif model == 'S12500' or model == 'S7600':
        res = parse_port_H3CS12500(port)
        if res != None:  # 返回不是NONE，说明端口格式正确, 则 定位到该 Port Obj
            portObj = Port.objects.filter(asset=asset,
                                          speed=res[0],
                                          chassisNum=res[1],
                                          slotNum=res[2],
                                          subSlotNum=res[3],
                                          portNum=res[4])[0]

        else:  # 返回值为NONE，说明端口格式有误
            return None

    elif model == 'S9300':
        res = parse_port_HWS9300(port)
        if res != None:  # 返回不是NONE，说明端口格式正确, 则 定位到该 Port Obj
            portObj = Port.objects.filter(asset=asset,
                                          speed=res[0],
                                          chassisNum=res[1],
                                          slotNum=res[2],
                                          subSlotNum=res[3],
                                          portNum=res[4])[0]

        else:  # 返回值为NONE，说明端口格式有误
            return None


    elif model == 'Brocade 6510':
        res = parse_port_Brocade6510(port)
        if res != None:  # 返回不是NONE，说明端口格式正确, 则 定位到该 Port Obj
            portObj = Port.objects.filter(asset=asset,
                                          slotNum=res[0],
                                          portNum=res[1])[0]

        else:  # 返回值为NONE，说明端口格式有误
            return None


    elif model == 'E8000E-X8':
        res = parse_port_E8000E(port)
        if res != None:  # 返回不是NONE，说明端口格式正确, 则 定位到该 Port Obj
            portObj = Port.objects.filter(asset=asset,
                                          speed=res[0],
                                          slotNum=res[1],
                                          subSlotNum=res[2],
                                          portNum=res[3])[0]

        else:  # 返回值为NONE，说明端口格式有误
            return None

    elif asset.asset_type == 'server' or asset.asset_type == 'storagedevice':  # 如果是服务器或者存储设备

        portObj = Port.objects.filter(asset=asset,
                                      portNum=port)[0]

    else:     # 如果是没有定义 parse_port方法的设备，则按照服务器来定位到port Obj
        portObj = Port.objects.filter(asset=asset,
                                      portNum=port)[0]


    return portObj


def assignPortToPort(tableFilePath):
    wb = load_workbook(tableFilePath)

    # 打开第一个sheet
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])

    err_portMutiAssign_log = []

    for cell in sheet['B']:
        if cell.value != None and cell.value.startswith("ZJ"): # 如果这一是 “有料的”
            A_port = get_portObj(sheet, 'B', 'D', cell.row)
            Z_port = get_portObj(sheet, 'E', 'G', cell.row)
            if A_port!=None and Z_port!=None:
                if A_port.remotePort!=None : # 如果A端口 已经存在关联
                    log = "第{0}行 ：A端<{1}>  已经连线到  Z端<{2}>"\
                        .format(str(cell.row), A_port, A_port.remotePort)
                    err_portMutiAssign_log.append(log)

                if hasattr(A_port, 'port'): # 如果A端口 已经 被关联
                    log = "第{0}行 ：Z端<{1}>  已经连线到  A端<{2}>"\
                        .format(str(cell.row), A_port.port, A_port)
                    err_portMutiAssign_log.append(log)

                if Z_port.remotePort != None:  # 如果Z端口 已经存在关联
                    log = "第{0}行 ： Z端<{1}>  已经连线到  A端<{2}>"\
                        .format(str(cell.row), Z_port, Z_port.remotePort)
                    err_portMutiAssign_log.append(log)

                if hasattr(Z_port, 'port'):  # 如果Z端口 已经 被关联
                    log = "第{0}行 ：A端<{1}>  已经连线到  Z端<{2}>" \
                        .format(str(cell.row), Z_port.port, Z_port)
                    err_portMutiAssign_log.append(log)

                elif A_port.remotePort==None and Z_port.remotePort==None and hasattr(A_port, 'port')==False and hasattr(Z_port, 'port')==False:# 如果A端口、Z端口 都没有关联别的端口 也都没有被别的端口关联
                    A_port.remotePort = Z_port
                    A_port.save()

    if len(err_portMutiAssign_log)==0:
        err_portMutiAssign_log.append('不存在冲突的布线！')

    return err_portMutiAssign_log



def check_samePort_on_stackDev(project):
    """
    传入 一个project Obj
    获取这个项目内的堆叠设备的端口，检查是否有命名重复的端口
    :param project:
    :return:
    """
    asset_Set = Asset.objects.filter(project=project)
    stackDevName_Set = set([])

    warn_samePort_log = []

    for asset_Obj in asset_Set:
        if len(Asset.objects.filter(sysname=asset_Obj.sysname))==2:# 如果是堆叠设备
            stackDevName_Set.add(asset_Obj.sysname)

    for stackDevName in stackDevName_Set:



        A_dev_portSet =  Asset.objects.filter(sysname=stackDevName)[0].port_set.all()
        B_dev_portSet = Asset.objects.filter(sysname=stackDevName)[1].port_set.all()

        A_dev_portNameSet = set([])
        for port in A_dev_portSet:
            A_dev_portNameSet.add(Port.__str__(port))

        for port in B_dev_portSet:
            portname = Port.__str__(port)
            if portname in A_dev_portNameSet:
                log = "堆叠设备 <{0}> 与 <{1}> 有相同的端口: {2}"\
                    .format(Asset.objects.filter(sysname=stackDevName)[0].name,
                            Asset.objects.filter(sysname=stackDevName)[1].name,
                            portname)
                if port.portNum.upper()!='MGMT':
                    warn_samePort_log.append(log)


    if len(warn_samePort_log)==0:
        warn_samePort_log.append('堆叠设备的端口命名均正确！')

    return warn_samePort_log






















